"""
Authentication and user management endpoints.

Login uses JWT. Invites create a user without a password and hand back a
signed set-password link. In production you would email that link, for the
demo it is returned to the admin so they can copy it.
"""
from datetime import timedelta
import random

from django.conf import settings
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework_simplejwt.tokens import RefreshToken

from integrations.email import send_invite_email, send_otp_email
from notifications.utils import ensure_default_preferences

from .models import Hospital, InviteStatus, PasswordResetOTP, Roles, User
from .permissions import IsAdmin, IsAdminOrCS
from .serializers import HospitalSerializer, UserSerializer
from .tokens import make_invite_token, read_invite_token


def tokens_for(user):
    refresh = RefreshToken.for_user(user)
    return {"access": str(refresh.access_token), "refresh": str(refresh)}


@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    """Sign in with email and password. Returns JWT tokens plus the profile."""
    email = (request.data.get("email") or "").strip().lower()
    password = request.data.get("password") or ""
    try:
        user = User.objects.get(email__iexact=email)
    except User.DoesNotExist:
        return Response({"detail": "Invalid email or password."}, status=status.HTTP_401_UNAUTHORIZED)
    if not user.check_password(password):
        return Response({"detail": "Invalid email or password."}, status=status.HTTP_401_UNAUTHORIZED)
    if user.invite_status == InviteStatus.DEACTIVATED:
        return Response({"detail": "This account has been deactivated. Contact your administrator."}, status=status.HTTP_403_FORBIDDEN)
    if user.invite_status == InviteStatus.INVITED:
        user.invite_status = InviteStatus.ACTIVE
        user.save(update_fields=["invite_status"])
    return Response({**tokens_for(user), "user": UserSerializer(user).data})


@api_view(["GET", "PATCH"])
def me_view(request):
    """Read or update the signed in user's own profile."""
    user = request.user
    if request.method == "PATCH":
        serializer = UserSerializer(user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
    return Response(UserSerializer(user).data)


@api_view(["POST"])
@permission_classes([AllowAny])
def set_password_view(request):
    """Activate an invited account, or reset a password, using a signed token."""
    token = request.data.get("token") or ""
    password = request.data.get("password") or ""
    if len(password) < 8:
        return Response({"detail": "Password must be at least 8 characters."}, status=400)
    uid = read_invite_token(token)
    if uid is None:
        return Response({"detail": "This link is invalid or has expired. Ask your administrator to resend it."}, status=400)
    try:
        user = User.objects.get(id=uid)
    except User.DoesNotExist:
        return Response({"detail": "Account not found."}, status=400)
    user.set_password(password)
    if user.invite_status == InviteStatus.INVITED:
        user.invite_status = InviteStatus.ACTIVE
    user.save()
    return Response({**tokens_for(user), "user": UserSerializer(user).data})


@api_view(["GET", "POST"])
@permission_classes([IsAdmin])
def users_view(request):
    """Admin: list every account, or invite a new one."""
    if request.method == "GET":
        users = User.objects.select_related("hospital", "manager").prefetch_related("programs").order_by("-created_at")
        return Response(UserSerializer(users, many=True).data)

    email = (request.data.get("email") or "").strip().lower()
    full_name = (request.data.get("full_name") or "").strip()
    role = request.data.get("role") or Roles.CLIENT
    if not email or not full_name:
        return Response({"detail": "Email and full name are required."}, status=400)
    if role not in Roles.values:
        return Response({"detail": "Unknown role."}, status=400)
    if User.objects.filter(email__iexact=email).exists():
        return Response({"detail": "A user with this email already exists."}, status=400)

    user = User.objects.create_user(
        email=email,
        full_name=full_name,
        role=role,
        hospital_id=request.data.get("hospital") or None,
        manager_id=request.data.get("manager") or None,
        invited_by=request.user,
    )
    ensure_default_preferences(user)
    invite_link = f"{settings.FRONTEND_URL}/set-password?token={make_invite_token(user.id)}"
    email_result = send_invite_email(user.email, user.full_name, invite_link)
    # The invite link is always returned too, so the admin can copy it
    # directly if the email did not go through (or SendGrid is not set up).
    return Response({"user": UserSerializer(user).data, "invite_link": invite_link, "email": email_result}, status=201)


@api_view(["PATCH"])
@permission_classes([IsAdmin])
def user_detail_view(request, user_id):
    """Admin: update a user's role, hospital, manager, programs, or status."""
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)

    for field in ["full_name", "phone"]:
        if field in request.data:
            setattr(user, field, request.data[field])
    if "role" in request.data and request.data["role"] in Roles.values:
        user.role = request.data["role"]
    if "hospital" in request.data:
        user.hospital_id = request.data["hospital"] or None
    if "manager" in request.data:
        user.manager_id = request.data["manager"] or None
    if "invite_status" in request.data and request.data["invite_status"] in InviteStatus.values:
        user.invite_status = request.data["invite_status"]
    user.save()
    if "program_ids" in request.data:
        user.programs.set(request.data["program_ids"] or [])
    return Response(UserSerializer(user).data)


@api_view(["POST"])
@permission_classes([IsAdmin])
def resend_invite_view(request, user_id):
    """Admin: generate a fresh set-password link for an invited user."""
    try:
        user = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return Response({"detail": "User not found."}, status=404)
    invite_link = f"{settings.FRONTEND_URL}/set-password?token={make_invite_token(user.id)}"
    email_result = send_invite_email(user.email, user.full_name, invite_link)
    return Response({"invite_link": invite_link, "email": email_result})


@api_view(["GET", "POST"])
@permission_classes([IsAdminOrCS])
def hospitals_view(request):
    """List hospitals, or add one (admin and customer service)."""
    if request.method == "POST":
        name = (request.data.get("name") or "").strip()
        if not name:
            return Response({"detail": "Hospital name is required."}, status=400)
        hospital = Hospital.objects.create(name=name)
        return Response(HospitalSerializer(hospital).data, status=201)
    return Response(HospitalSerializer(Hospital.objects.order_by("name"), many=True).data)


@api_view(["GET"])
def staff_directory_view(request):
    """
    Employee directory for the Customer Service view.

    Supports the employee search bar and the program filter:
      ?q=nia            search by name or email
      ?role=field_staff  filter by role
      ?program=3         filter by program id
      ?sort=program      sort by first program name instead of full name
    """
    if request.user.role not in ("admin", "customer_service", "manager"):
        return Response({"detail": "Not allowed."}, status=403)

    qs = (
        User.objects.filter(role__in=["field_staff", "customer_service", "manager"])
        .prefetch_related("programs")
        .select_related("manager")
    )
    q = request.query_params.get("q", "").strip()
    if q:
        from django.db.models import Q
        qs = qs.filter(Q(full_name__icontains=q) | Q(email__icontains=q))
    role = request.query_params.get("role", "").strip()
    if role:
        qs = qs.filter(role=role)
    program = request.query_params.get("program", "").strip()
    if program:
        qs = qs.filter(programs__id=program)
    if request.query_params.get("sort") == "program":
        qs = qs.order_by("programs__name", "full_name")
    else:
        qs = qs.order_by("full_name")
    return Response(UserSerializer(qs.distinct(), many=True).data)


@api_view(["POST"])
@permission_classes([IsAdmin])
def bulk_invite_view(request):
    """
    Admin: upload a spreadsheet of new users and invite all of them at once.

    Expected columns, header row required, order does not matter:
      full_name, email, role, programs, manager_email

    - role must be one of the CareLink roles (admin, manager,
      customer_service, field_staff, hospital_partner, client, family).
    - programs is optional, a comma separated list of existing program
      names. Names that do not match an existing program are skipped and
      reported back, they are not created here (see the Programs bulk
      upload for that).
    - manager_email is optional, used for field_staff rows to set who
      approves their shift changes. The manager must already exist.

    Every row that succeeds gets a real invite email sent through SendGrid
    (or printed to the console if SendGrid is not configured).
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Upload an .xlsx file with a file field named file."}, status=400)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return Response({"detail": "openpyxl is not installed on the server."}, status=500)

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return Response({"detail": "Could not read that file. Make sure it is a valid .xlsx spreadsheet."}, status=400)

    if not rows:
        return Response({"detail": "The spreadsheet is empty."}, status=400)

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    required = ["full_name", "email", "role"]
    if not all(col in header for col in required):
        return Response({"detail": f"Header row must include: {', '.join(required)} (programs and manager_email are optional)."}, status=400)

    col = {name: header.index(name) for name in header if name}

    created, skipped, errors = [], [], []
    from care.models import Program

    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        full_name = str(row[col["full_name"]] or "").strip() if col.get("full_name") is not None else ""
        email = str(row[col["email"]] or "").strip().lower() if col.get("email") is not None else ""
        role = str(row[col["role"]] or "").strip().lower() if col.get("role") is not None else ""
        programs_raw = str(row[col["programs"]] or "").strip() if "programs" in col and col["programs"] is not None and len(row) > col["programs"] else ""
        manager_email = str(row[col["manager_email"]] or "").strip().lower() if "manager_email" in col and col["manager_email"] is not None and len(row) > col["manager_email"] else ""

        if not full_name or not email or not role:
            errors.append({"row": row_number, "reason": "Missing full_name, email, or role."})
            continue
        if role not in Roles.values:
            errors.append({"row": row_number, "email": email, "reason": f"Unknown role '{role}'."})
            continue
        if User.objects.filter(email__iexact=email).exists():
            skipped.append({"row": row_number, "email": email, "reason": "A user with this email already exists."})
            continue

        manager = None
        if manager_email:
            manager = User.objects.filter(email__iexact=manager_email, role=Roles.MANAGER).first()
            if manager is None:
                errors.append({"row": row_number, "email": email, "reason": f"Manager '{manager_email}' not found."})
                continue

        user = User.objects.create_user(email=email, full_name=full_name, role=role, manager=manager, invited_by=request.user)
        ensure_default_preferences(user)

        if programs_raw:
            names = [p.strip() for p in programs_raw.split(",") if p.strip()]
            matched = Program.objects.filter(name__in=names)
            user.programs.set(matched)
            missing = set(names) - set(matched.values_list("name", flat=True))
            if missing:
                errors.append({"row": row_number, "email": email, "reason": f"Programs not found, skipped: {', '.join(missing)}"})

        invite_link = f"{settings.FRONTEND_URL}/set-password?token={make_invite_token(user.id)}"
        send_invite_email(user.email, user.full_name, invite_link)
        created.append({"row": row_number, "email": email, "role": role})

    return Response({"created": created, "skipped": skipped, "errors": errors})


@api_view(["POST"])
@permission_classes([AllowAny])
def request_password_reset_view(request):
    """Email a 6 digit code that expires in 10 minutes. Always returns a
    generic success message, whether or not the email matches an account,
    so this endpoint cannot be used to discover who has a CareLink account."""
    email = (request.data.get("email") or "").strip().lower()
    user = User.objects.filter(email__iexact=email).first()
    if user is not None:
        code = f"{random.randint(0, 999999):06d}"
        PasswordResetOTP.objects.create(user=user, code=code, expires_at=timezone.now() + timedelta(minutes=10))
        send_otp_email(user.email, code)
    return Response({"detail": "If that email has a CareLink account, a reset code has been sent."})


@api_view(["POST"])
@permission_classes([AllowAny])
def verify_password_reset_view(request):
    """Check the code and, if valid, set the new password and sign the user in."""
    email = (request.data.get("email") or "").strip().lower()
    code = (request.data.get("code") or "").strip()
    password = request.data.get("password") or ""
    if len(password) < 8:
        return Response({"detail": "Password must be at least 8 characters."}, status=400)

    user = User.objects.filter(email__iexact=email).first()
    if user is None:
        return Response({"detail": "Invalid or expired code."}, status=400)

    otp = PasswordResetOTP.objects.filter(user=user, code=code, used=False, expires_at__gte=timezone.now()).order_by("-created_at").first()
    if otp is None:
        return Response({"detail": "Invalid or expired code."}, status=400)

    otp.used = True
    otp.save(update_fields=["used"])
    user.set_password(password)
    if user.invite_status == InviteStatus.INVITED:
        user.invite_status = InviteStatus.ACTIVE
    user.save()
    return Response({**tokens_for(user), "user": UserSerializer(user).data})