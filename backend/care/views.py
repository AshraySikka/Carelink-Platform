"""
Care domain endpoints. Every list is scoped to what the caller's role is
allowed to see, mirroring the row level security the original Supabase
version enforced in the database.
"""
import math

from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response

from accounts.models import Roles, User
from accounts.permissions import IsAdmin, IsAdminOrCS, IsSchedulingStaff
from notifications.utils import notify, notify_role
from .models import (
    BLOCKED_REASON_CODES, ChangeReasonCode, ChangeRequestStatus, ChangeRequestType,
    ClinicalDocument, EmergencyRequest, FamilyMember, NewsPost, PlatformSetting,
    Program, Referral, ReferralDocument, ReferralStatus, Resource, Shift,
    ShiftChangeRequest, ShiftStatus,
)
from .serializers import (
    ClinicalDocumentSerializer, EmergencySerializer, FamilyMemberSerializer,
    NewsPostSerializer, ProgramSerializer, ReferralSerializer, ResourceSerializer,
    ShiftChangeRequestSerializer, ShiftSerializer,
)

from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer

from messaging.models import Conversation, ConversationParticipant, Message

# ---------------- Programs ----------------

@api_view(["GET", "POST"])
def programs_view(request):
    """List programs (any signed in staff role), create one (admin only)."""
    if request.method == "POST":
        if request.user.role != Roles.ADMIN:
            return Response({"detail": "Only admins can create programs."}, status=403)
        serializer = ProgramSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=201)
    return Response(ProgramSerializer(Program.objects.order_by("name"), many=True).data)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdmin])
def program_detail_view(request, program_id):
    try:
        program = Program.objects.get(id=program_id)
    except Program.DoesNotExist:
        return Response({"detail": "Program not found."}, status=404)
    if request.method == "DELETE":
        program.delete()
        return Response(status=204)
    serializer = ProgramSerializer(program, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


# ---------------- Referrals ----------------

def _referral_queryset_for(user):
    qs = Referral.objects.select_related("hospital", "submitted_by", "assigned_staff").prefetch_related("documents")
    if user.role in (Roles.ADMIN, Roles.CUSTOMER_SERVICE, Roles.MANAGER):
        return qs
    if user.role == Roles.HOSPITAL_PARTNER:
        return qs.filter(submitted_by=user)
    return qs.none()


@api_view(["GET", "POST"])
def referrals_view(request):
    if request.method == "GET":
        return Response(ReferralSerializer(_referral_queryset_for(request.user).order_by("-created_at"), many=True).data)

    # Only hospital partners with an org affiliation can submit referrals.
    if request.user.role != Roles.HOSPITAL_PARTNER or not request.user.hospital_id:
        return Response({"detail": "Only hospital partners with an organization can submit referrals."}, status=403)
    serializer = ReferralSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    referral = serializer.save(hospital=request.user.hospital, submitted_by=request.user, source="portal")
    # New referral notification goes to the customer service team.
    notify_role(Roles.CUSTOMER_SERVICE, "referrals", "New referral received",
                f"{referral.client_name} was referred by {request.user.hospital.name}.", link="/cs/queue")
    return Response(ReferralSerializer(referral).data, status=201)


@api_view(["PATCH"])
def referral_detail_view(request, referral_id):
    try:
        referral = _referral_queryset_for(request.user).get(id=referral_id)
    except Referral.DoesNotExist:
        return Response({"detail": "Referral not found."}, status=404)

    # Hospital partners may only edit their own referral while it is still new.
    if request.user.role == Roles.HOSPITAL_PARTNER and referral.status != ReferralStatus.NEW:
        return Response({"detail": "This referral is read only once accepted."}, status=403)

    old_status = referral.status
    serializer = ReferralSerializer(referral, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    referral = serializer.save()
    if referral.status != old_status:
        notify(referral.submitted_by, "referrals", "Referral status updated",
               f"{referral.client_name} is now {referral.get_status_display()}.", link="/hospital")
    return Response(ReferralSerializer(referral).data)


@api_view(["POST"])
def referral_documents_view(request, referral_id):
    """
    Attach an uploaded file to a referral. Not restricted to any particular
    referral status, so a hospital partner can add documents at any point,
    not just while the referral is still new.
    """
    try:
        referral = _referral_queryset_for(request.user).get(id=referral_id)
    except Referral.DoesNotExist:
        return Response({"detail": "Referral not found."}, status=404)
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No file provided."}, status=400)
    if file.size > 10 * 1024 * 1024:
        return Response({"detail": "Files must be 10MB or smaller."}, status=400)
    doc = ReferralDocument.objects.create(referral=referral, uploaded_by=request.user, file=file, file_name=file.name)
    return Response({"id": doc.id, "file_name": doc.file_name, "file": doc.file.url}, status=201)



# ---------------- Shifts ----------------

def _shift_queryset_for(user):
    qs = Shift.objects.select_related("field_staff", "client")
    if user.role in (Roles.ADMIN, Roles.CUSTOMER_SERVICE, Roles.MANAGER):
        return qs
    if user.role == Roles.FIELD_STAFF:
        return qs.filter(field_staff=user)
    if user.role == Roles.CLIENT:
        return qs.filter(client=user)
    if user.role == Roles.FAMILY:
        client_ids = FamilyMember.objects.filter(family_user=user).values_list("client_id", flat=True)
        return qs.filter(client_id__in=client_ids)
    return qs.none()


@api_view(["GET", "POST"])
def shifts_view(request):
    if request.method == "GET":
        return Response(ShiftSerializer(_shift_queryset_for(request.user).order_by("start_time"), many=True).data)
    if request.user.role not in (Roles.ADMIN, Roles.CUSTOMER_SERVICE, Roles.MANAGER):
        return Response({"detail": "Only scheduling staff can create shifts."}, status=403)
    serializer = ShiftSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    shift = serializer.save()
    notify(shift.field_staff, "schedule", "New shift assigned",
           f"Visit with {shift.client.full_name} on {shift.start_time:%b %d, %I:%M %p}.", link="/field")
    notify(shift.client, "schedule", "New visit scheduled",
           f"{shift.field_staff.full_name} will visit on {shift.start_time:%b %d, %I:%M %p}.", link="/care")
    return Response(ShiftSerializer(shift).data, status=201)


@api_view(["PATCH", "DELETE"])
def shift_detail_view(request, shift_id):
    try:
        shift = _shift_queryset_for(request.user).get(id=shift_id)
    except Shift.DoesNotExist:
        return Response({"detail": "Shift not found."}, status=404)
    if request.method == "DELETE":
        if request.user.role not in (Roles.ADMIN, Roles.CUSTOMER_SERVICE):
            return Response({"detail": "Only scheduling staff can delete shifts."}, status=403)
        shift.delete()
        return Response(status=204)
    old_start, old_end, old_status = shift.start_time, shift.end_time, shift.status
    serializer = ShiftSerializer(shift, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    shift = serializer.save()
    # Manually moving a shift off "change requested" clears the stale
    # pending-change fields, so the schedule stops showing a request that
    # has already been handled by hand.
    if request.data.get("status") and request.data["status"] != ShiftStatus.CHANGE_REQUESTED and shift.change_request_note:
        shift.change_request_note = ""
        shift.requested_start_time = None
        shift.requested_end_time = None
        shift.save(update_fields=["change_request_note", "requested_start_time", "requested_end_time"])
    # A real time change gets a notification, which also doubles as the
    # signal any open field staff or client screen uses to refresh itself
    # live instead of showing a stale time until the next page load.
    if shift.start_time != old_start or shift.end_time != old_end:
        new_range = f"{timezone.localtime(shift.start_time):%b %d, %I:%M %p} to {timezone.localtime(shift.end_time):%I:%M %p}"
        notify(shift.field_staff, "schedule", "Your shift time was updated", f"New time: {new_range}.", link="/field")
        notify(shift.client, "schedule", "Your visit time was updated", f"New time: {new_range}.", link="/care")
    if request.data.get("status") == ShiftStatus.CANCELLED and old_status != ShiftStatus.CANCELLED:
        notify(shift.field_staff, "schedule", "Visit cancelled",
               f"Your visit with {shift.client.full_name} on {shift.start_time:%b %d, %I:%M %p} was cancelled.", link="/field")
        notify(shift.client, "schedule", "Your visit was cancelled",
               f"Your visit on {shift.start_time:%b %d, %I:%M %p} was cancelled.", link="/care")
    return Response(ShiftSerializer(shift).data)


def _haversine_meters(lat1, lng1, lat2, lng2):
    """Distance between two coordinates in meters, used by the clock in geofence."""
    radius = 6371000
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dp, dl = math.radians(lat2 - lat1), math.radians(lng2 - lng1)
    a = math.sin(dp / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dl / 2) ** 2
    return 2 * radius * math.asin(math.sqrt(a))


@api_view(["POST"])
def shift_clock_in_view(request, shift_id):
    """
    Clock in, gated to 15 minutes before the start and 100m of the client
    address. Clocking in from farther away requires a written reason: it
    gets logged on the shift, sent to the field staff member's manager,
    and shows up in the "Clock-in location overrides" report so patterns
    (the same excuse every time, the same staff member every time) are
    visible rather than buried in individual notifications.
    """
    try:
        shift = Shift.objects.select_related("client", "field_staff", "field_staff__manager").get(id=shift_id, field_staff=request.user)
    except Shift.DoesNotExist:
        return Response({"detail": "Shift not found."}, status=404)
    if shift.clock_in_at:
        return Response({"detail": "Already clocked in."}, status=400)
    now = timezone.now()
    if now < shift.start_time - timezone.timedelta(minutes=15):
        return Response({"detail": "You can only clock in within 15 minutes of the scheduled start."}, status=400)

    override = False
    override_distance = None
    override_reason = (request.data.get("override_reason") or "").strip()
    lat, lng = request.data.get("latitude"), request.data.get("longitude")
    client = shift.client
    if lat is not None and lng is not None and client.latitude is not None and client.longitude is not None:
        distance = _haversine_meters(float(lat), float(lng), client.latitude, client.longitude)
        if distance > 100:
            if not request.data.get("override"):
                return Response(
                    {"detail": f"You are {round(distance)}m from the client address (limit 100m). Confirm override to clock in anyway.", "needs_override": True},
                    status=400,
                )
            if not override_reason:
                return Response(
                    {"detail": "Tell us why you're outside the client's location before we can log this.", "needs_override_reason": True},
                    status=400,
                )
            override = True
            override_distance = round(distance)

    shift.clock_in_at = now
    shift.status = ShiftStatus.IN_PROGRESS
    shift.geofence_override = override
    shift.geofence_override_reason = override_reason if override else ""
    shift.geofence_override_distance_meters = override_distance
    shift.save()

    if override:
        distance_text = f"{override_distance}m" if override_distance is not None else "an unrecorded distance"
        detail = f"{request.user.full_name} clocked in {distance_text} from {client.full_name}'s address. Reason: {override_reason}"
        if shift.field_staff.manager:
            notify(shift.field_staff.manager, "schedule", "Clock-in location override", detail, link="/cs/schedule")

    return Response(ShiftSerializer(shift).data)


# Staff cannot clock out until this many minutes before the shift's
# scheduled end time. No override, unlike the clock-in geofence check: this
# exists specifically to stop an early checkout, so it stays a hard block.
CLOCK_OUT_EARLY_MINUTES = 7


@api_view(["POST"])
def shift_clock_out_view(request, shift_id):
    try:
        shift = Shift.objects.get(id=shift_id, field_staff=request.user)
    except Shift.DoesNotExist:
        return Response({"detail": "Shift not found."}, status=404)
    if not shift.clock_in_at:
        return Response({"detail": "You need to clock in before you can clock out."}, status=400)
    if shift.clock_out_at:
        return Response({"detail": "Already clocked out."}, status=400)
    now = timezone.now()
    earliest = shift.end_time - timezone.timedelta(minutes=CLOCK_OUT_EARLY_MINUTES)
    if now < earliest:
        minutes_left = math.ceil((earliest - now).total_seconds() / 60)
        return Response(
            {"detail": f"You can clock out starting {CLOCK_OUT_EARLY_MINUTES} minutes before the shift ends. {minutes_left} more minute(s) to go."},
            status=400,
        )
    shift.clock_out_at = now
    shift.status = ShiftStatus.COMPLETED
    shift.save()
    return Response(ShiftSerializer(shift).data)


@api_view(["POST"])
def shift_on_my_way_view(request, shift_id):
    """
    Sends the client a heads up that the caregiver is on the way, and also
    posts that as a real chat message so it shows up in Messages, not just
    a notification. Only allowed on the calendar day of the shift itself,
    so this cannot be sent a week early by mistake.
    """
    try:
        shift = Shift.objects.select_related("client").get(id=shift_id, field_staff=request.user)
    except Shift.DoesNotExist:
        return Response({"detail": "Shift not found."}, status=404)
    if shift.on_my_way_at:
        return Response({"detail": "You already sent this for this visit."}, status=400)
    if timezone.localtime(shift.start_time).date() != timezone.localdate():
        return Response({"detail": "On my way can only be sent on the day of the shift."}, status=400)

    shift.on_my_way_at = timezone.now()
    shift.save()
    notify(shift.client, "schedule", "Your caregiver is on the way",
           f"{request.user.full_name} is heading to your visit now.", link="/care")

    # Post the same update as a real chat message, reusing or creating the
    # direct conversation between this field staff member and the client.

    conversation = (
        Conversation.objects.filter(participants__user=request.user)
        .filter(participants__user=shift.client)
        .first()
    )
    if conversation is None:
        conversation = Conversation.objects.create(created_by=request.user)
        ConversationParticipant.objects.create(conversation=conversation, user=request.user)
        ConversationParticipant.objects.create(conversation=conversation, user=shift.client)

    message = Message.objects.create(conversation=conversation, sender=request.user, body="On my way!")
    payload = {
        "kind": "message", "conversation_id": conversation.id, "id": message.id,
        "sender_id": request.user.id, "sender_name": request.user.full_name,
        "body": message.body, "created_at": message.created_at.isoformat(),
    }
    layer = get_channel_layer()
    for participant in ConversationParticipant.objects.filter(conversation=conversation).select_related("user"):
        try:
            async_to_sync(layer.group_send)(f"user_{participant.user_id}", {"type": "push", "payload": payload})
        except Exception:
            pass
        if participant.user_id != request.user.id:
            notify(participant.user, "messages", f"New message from {request.user.full_name}", message.body, link="/messages")

    return Response({**ShiftSerializer(shift).data, "conversation_id": conversation.id})


# ---------------- Shift change approval workflow ----------------

@api_view(["GET", "POST"])
def change_requests_view(request):
    """
    GET: the requests the caller should see.
      Managers see their own approval queue. Field staff see their requests.
      Admin and CS see everything.
    POST: field staff file a new request (reschedule or cancel). Their
      manager is notified. A blocked reason code (an in-progress safety or
      medical concern) is rejected here rather than queued, since those
      need customer service alerted immediately, not a manager decision
      that might sit unanswered.
    """
    user = request.user
    if request.method == "GET":
        qs = ShiftChangeRequest.objects.select_related("shift", "requested_by", "shift__client", "shift__field_staff")
        if user.role == Roles.MANAGER:
            qs = qs.filter(manager=user)
        elif user.role == Roles.FIELD_STAFF:
            qs = qs.filter(requested_by=user)
        elif user.role == Roles.CLIENT:
            qs = qs.filter(requested_by=user)
        elif user.role not in (Roles.ADMIN, Roles.CUSTOMER_SERVICE):
            qs = qs.none()
        return Response(ShiftChangeRequestSerializer(qs.order_by("-created_at"), many=True).data)

    # Filing a request. Field staff go through manager approval. Clients can
    # also request changes, those route straight to customer service.
    reason_code = request.data.get("reason_code")
    if reason_code not in ChangeReasonCode.values:
        return Response({"detail": "Pick a reason from the list."}, status=400)
    if reason_code in BLOCKED_REASON_CODES:
        return Response({
            "detail": "This needs immediate attention, not a queued request. Please use Emergency request instead so customer service is alerted right away.",
            "redirect": "emergency",
        }, status=400)
    request_type = request.data.get("request_type")
    if request_type not in ChangeRequestType.values:
        return Response({"detail": "Pick reschedule or cancel."}, status=400)

    other_text = (request.data.get("reason_other") or "").strip()
    label = ChangeReasonCode(reason_code).label
    type_label = "Cancellation" if request_type == ChangeRequestType.CANCEL else "Reschedule"
    reason_text = f"{type_label} - {label}"
    if reason_code == ChangeReasonCode.OTHER and other_text:
        reason_text += f": {other_text}"
    elif other_text:
        reason_text += f" ({other_text})"

    payload = {
        "shift": request.data.get("shift"),
        "request_type": request_type,
        "reason_code": reason_code,
        "reason": reason_text,
        "requested_start_time": request.data.get("requested_start_time"),
        "requested_end_time": request.data.get("requested_end_time"),
    }
    serializer = ShiftChangeRequestSerializer(data=payload)
    serializer.is_valid(raise_exception=True)
    shift = serializer.validated_data["shift"]

    if user.role == Roles.FIELD_STAFF:
        if shift.field_staff_id != user.id:
            return Response({"detail": "You can only request changes to your own shifts."}, status=403)
        if not user.manager_id:
            return Response({"detail": "You do not have a manager assigned yet. Ask your administrator to set one."}, status=400)
        change = serializer.save(requested_by=user, manager=user.manager)
        shift.status = ShiftStatus.CHANGE_REQUESTED
        shift.change_request_note = change.reason
        shift.requested_start_time = change.requested_start_time
        shift.requested_end_time = change.requested_end_time
        shift.save()
        notify(user.manager, "approvals", f"{type_label} request needs your approval",
               f"{user.full_name} requested a {type_label.lower()} for their {shift.start_time:%b %d} shift: {change.reason}", link="/approvals")
        return Response(ShiftChangeRequestSerializer(change).data, status=201)

    if user.role == Roles.CLIENT:
        if shift.client_id != user.id:
            return Response({"detail": "You can only request changes to your own visits."}, status=403)
        change = serializer.save(requested_by=user, manager=None, status=ChangeRequestStatus.APPROVED)
        shift.status = ShiftStatus.CHANGE_REQUESTED
        shift.change_request_note = change.reason
        shift.requested_start_time = change.requested_start_time
        shift.requested_end_time = change.requested_end_time
        shift.save()
        notify_role(Roles.CUSTOMER_SERVICE, "schedule", f"Client requested a {type_label.lower()}",
                    f"{user.full_name}: {change.reason}", link="/cs/schedule")
        notify(shift.field_staff, "schedule", f"Client requested a {type_label.lower()} to your visit",
               f"{user.full_name}: {change.reason}", link="/field")
        return Response(ShiftChangeRequestSerializer(change).data, status=201)

    return Response({"detail": "Only field staff and clients file change requests."}, status=403)


@api_view(["POST"])
def change_request_decide_view(request, request_id):
    """
    Manager, customer service, or admin approves or declines.

    Customer service and admins can decide a request too, not just the
    assigned manager, so an unanswered request is never stuck waiting on
    one person who might be unreachable. See escalate_change_requests for
    the scheduled job that pings everyone again as the shift gets close.

    Approving does NOT change the shift's actual time by itself. It only
    marks the request approved and tells customer service exactly what was
    asked for, so a real person applies it deliberately on the Change
    requests screen. Until that happens, the shift keeps its original time
    and stays under the field staff member's upcoming visits, it does not
    disappear or show a stale pending badge. Same day requests are flagged
    as high priority since there is no time to spare.
    """
    try:
        change = ShiftChangeRequest.objects.select_related("shift", "requested_by", "shift__client").get(id=request_id)
    except ShiftChangeRequest.DoesNotExist:
        return Response({"detail": "Request not found."}, status=404)
    user = request.user
    if user.role not in (Roles.ADMIN, Roles.CUSTOMER_SERVICE) and change.manager_id != user.id:
        return Response({"detail": "Only the assigned manager, customer service, or an admin can decide this request."}, status=403)
    if change.status != ChangeRequestStatus.PENDING:
        return Response({"detail": "This request has already been decided."}, status=400)

    decision = request.data.get("decision")
    if decision not in ("approved", "declined"):
        return Response({"detail": "Decision must be approved or declined."}, status=400)

    shift = change.shift
    is_same_day = timezone.localtime(shift.start_time).date() == timezone.localdate()
    priority = "High priority, same day: " if is_same_day else ""

    requested_range = None
    if change.requested_start_time:
        requested_range = f"{timezone.localtime(change.requested_start_time):%b %d, %I:%M %p}"
        if change.requested_end_time:
            requested_range += f" to {timezone.localtime(change.requested_end_time):%I:%M %p}"

    change.status = decision
    change.decided_by = user
    change.decision_note = request.data.get("note", "")
    change.decided_at = timezone.now()
    change.save()

    # Approved: the shift shows a distinct "approved, pending change" status
    # so it is obviously not just a normal shift, until customer service
    # actually applies a new time. Declined: it goes straight back to normal.
    shift.status = ShiftStatus.APPROVED_PENDING_CHANGE if decision == "approved" else ShiftStatus.SCHEDULED
    shift.change_request_note = ""
    shift.requested_start_time = None
    shift.requested_end_time = None
    shift.save()

    if decision == "approved":
        detail = f"Reason: {change.reason}. "
        detail += f"Requested new time: {requested_range}. " if requested_range else "No specific new time was requested, follow up with the staff member. "
        detail += f"Approved by {user.full_name}. Update the shift on the Change requests screen."

        notify_role(Roles.CUSTOMER_SERVICE, "approvals", f"{priority}Shift change approved for {change.requested_by.full_name}",
                    detail, link="/cs/change-requests")
        notify(change.requested_by, "approvals", "Your shift change was approved",
               f"Approved by {user.full_name}. Customer service will update your schedule shortly.",
               link="/field")
    else:
        notify(change.requested_by, "approvals", "Your shift change was declined",
               f"Your request: {change.reason}. " + (change.decision_note or f"{user.full_name} declined this request. Talk to them for details."),
               link="/field")

    return Response(ShiftChangeRequestSerializer(change).data)


# ---------------- Emergencies ----------------

def _notify_family(client_user_id, category, title, body, link=""):
    """Notify every family member linked to a client. Used to keep the
    3 way flow (client, family, customer service) all in sync."""
    if not client_user_id:
        return
    for member in FamilyMember.objects.filter(client_id=client_user_id, family_user__isnull=False).select_related("family_user"):
        notify(member.family_user, category, title, body, link=link)


@api_view(["GET", "POST"])
def emergencies_view(request):
    user = request.user
    if request.method == "POST":
        description = (request.data.get("description") or "").strip()
        if not description:
            return Response({"detail": "Description is required."}, status=400)
        if user.role == Roles.CLIENT:
            emergency = EmergencyRequest.objects.create(client=user, source="client", description=description)
            # The client reported it themselves, so only their family needs
            # telling, not the client.
            _notify_family(user.id, "emergencies", "Emergency reported",
                            f"{user.full_name} reported: {description[:140]}", link="/family")
        else:
            emergency = EmergencyRequest.objects.create(reporter=user, source="staff", description=description,
                                                        client_id=request.data.get("client") or None)
            if emergency.client_id:
                notify(emergency.client, "emergencies", "An emergency was reported for your care",
                       description[:140], link="/care/emergencies")
                _notify_family(emergency.client_id, "emergencies", "An emergency was reported",
                                description[:140], link="/family")
        notify_role(Roles.CUSTOMER_SERVICE, "emergencies", "Emergency request",
                    description[:140], link="/cs/emergencies")
        return Response(EmergencySerializer(emergency).data, status=201)

    qs = EmergencyRequest.objects.select_related("client", "reporter").order_by("-created_at")
    if user.role in (Roles.ADMIN, Roles.CUSTOMER_SERVICE, Roles.MANAGER):
        pass
    elif user.role == Roles.CLIENT:
        qs = qs.filter(client=user)
    elif user.role == Roles.FIELD_STAFF:
        qs = qs.filter(reporter=user)
    elif user.role == Roles.FAMILY:
        client_ids = FamilyMember.objects.filter(family_user=user).values_list("client_id", flat=True)
        qs = qs.filter(client_id__in=client_ids)
    else:
        qs = qs.none()
    return Response(EmergencySerializer(qs, many=True).data)

@api_view(["PATCH"])
@permission_classes([IsAdminOrCS])
def emergency_detail_view(request, emergency_id):
    try:
        emergency = EmergencyRequest.objects.get(id=emergency_id)
    except EmergencyRequest.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    status_changed = "status" in request.data and request.data["status"] != emergency.status
    if "status" in request.data:
        emergency.status = request.data["status"]
    if "resolution_notes" in request.data:
        emergency.resolution_notes = request.data["resolution_notes"]
    emergency.save()

    # Closes the loop: the client and their family both hear about status
    # changes, not just customer service.
    if status_changed and emergency.client_id:
        label = {"new": "received", "acknowledged": "acknowledged", "resolved": "resolved"}.get(emergency.status, emergency.status)
        notify(emergency.client, "emergencies", f"Your emergency request was {label}",
               emergency.description[:140], link="/care/emergencies")
        _notify_family(emergency.client_id, "emergencies", f"Emergency request {label}",
                        emergency.description[:140], link="/family")
    if status_changed and emergency.reporter_id:
        # Whoever filed it (typically field staff) hears about the outcome
        # too, since it concerns them directly.
        label = {"new": "received", "acknowledged": "acknowledged", "resolved": "resolved"}.get(emergency.status, emergency.status)
        notify(emergency.reporter, "emergencies", f"Emergency you reported was {label}",
               emergency.description[:140], link="/field")
    return Response(EmergencySerializer(emergency).data)


# ---------------- Family ----------------

@api_view(["GET", "POST"])
def family_view(request):
    user = request.user
    if request.method == "POST":
        if user.role != Roles.CLIENT:
            return Response({"detail": "Only clients manage family access."}, status=403)
        serializer = FamilyMemberSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        member = serializer.save(client=user)
        # Auto link if a family account with this email already exists.
        existing = User.objects.filter(email__iexact=member.family_email, role=Roles.FAMILY).first()
        if existing:
            member.family_user = existing
            member.save()
        return Response(FamilyMemberSerializer(member).data, status=201)
    if user.role == Roles.CLIENT:
        qs = FamilyMember.objects.filter(client=user)
    elif user.role == Roles.FAMILY:
        qs = FamilyMember.objects.filter(family_user=user)
    else:
        qs = FamilyMember.objects.none()
    return Response(FamilyMemberSerializer(qs, many=True).data)


@api_view(["DELETE"])
def family_detail_view(request, member_id):
    try:
        member = FamilyMember.objects.get(id=member_id, client=request.user)
    except FamilyMember.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    member.delete()
    return Response(status=204)


# ---------------- Resources and news ----------------

@api_view(["GET", "POST"])
def resources_view(request):
    user = request.user
    if request.method == "POST":
        if user.role != Roles.ADMIN:
            return Response({"detail": "Only admins manage resources."}, status=403)
        serializer = ResourceSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=201)

    if user.role == Roles.ADMIN:
        # Admins manage the library, so they see every resource, published
        # or not, for every audience, not just what their own role can see.
        resources = Resource.objects.all().order_by("category", "title")
        return Response(ResourceSerializer(resources, many=True).data)

    # Everyone else only sees published resources targeted at their role.
    # An empty audience means everyone, same convention as news posts.
    resources = Resource.objects.filter(published=True).order_by("category", "title")
    visible = [r for r in resources if not r.audience or user.role in r.audience]
    return Response(ResourceSerializer(visible, many=True).data)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdmin])
def resource_detail_view(request, resource_id):
    try:
        resource = Resource.objects.get(id=resource_id)
    except Resource.DoesNotExist:
        return Response({"detail": "Resource not found."}, status=404)
    if request.method == "DELETE":
        resource.delete()
        return Response(status=204)
    serializer = ResourceSerializer(resource, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)



def _roles_at_cap(candidate_audience, exclude_id=None):
    """Return the list of roles that are already at the news post cap and
    would be pushed over it by publishing candidate_audience. An empty
    audience (everyone) is checked against every role."""
    setting = PlatformSetting.load()
    cap = setting.news_post_cap
    target_roles = list(Roles.values) if not candidate_audience else candidate_audience

    published = NewsPost.objects.filter(published=True)
    if exclude_id is not None:
        published = published.exclude(id=exclude_id)
    published = list(published)

    over = []
    for role in target_roles:
        count = sum(1 for p in published if not p.audience or role in p.audience)
        if count >= cap:
            over.append(role)
    return over


@api_view(["GET", "POST"])
def news_view(request):
    user = request.user
    if request.method == "POST":
        if user.role != Roles.ADMIN:
            return Response({"detail": "Only admins publish news."}, status=403)
        serializer = NewsPostSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        audience = serializer.validated_data.get("audience") or []
        if serializer.validated_data.get("published", True):
            over = _roles_at_cap(audience)
            if over:
                cap = PlatformSetting.load().news_post_cap
                return Response(
                    {"detail": f"These roles are already at the {cap} post limit, unpublish an older post first: {', '.join(over)}."},
                    status=400,
                )
        serializer.save(author=user)
        return Response(serializer.data, status=201)
    posts = NewsPost.objects.filter(published=True).order_by("-created_at")
    visible = [p for p in posts if not p.audience or user.role in p.audience]
    return Response(NewsPostSerializer(visible, many=True).data)


@api_view(["PATCH", "DELETE"])
@permission_classes([IsAdmin])
def news_detail_view(request, post_id):
    try:
        post = NewsPost.objects.get(id=post_id)
    except NewsPost.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.method == "DELETE":
        post.delete()
        return Response(status=204)
    serializer = NewsPostSerializer(post, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    will_publish = serializer.validated_data.get("published", post.published)
    audience = serializer.validated_data.get("audience", post.audience)
    if will_publish and (not post.published or audience != post.audience):
        over = _roles_at_cap(audience, exclude_id=post.id)
        if over:
            cap = PlatformSetting.load().news_post_cap
            return Response(
                {"detail": f"These roles are already at the {cap} post limit, unpublish an older post first: {', '.join(over)}."},
                status=400,
            )
    serializer.save()
    return Response(serializer.data)


@api_view(["GET", "PATCH"])
@permission_classes([IsAdmin])
def news_settings_view(request):
    """Admin: read or update the news post cap, clamped between 1 and 5."""
    setting = PlatformSetting.load()
    if request.method == "PATCH":
        try:
            cap = int(request.data.get("news_post_cap"))
        except (TypeError, ValueError):
            return Response({"detail": "news_post_cap must be a number."}, status=400)
        setting.news_post_cap = max(1, min(5, cap))
        setting.save()
    return Response({"news_post_cap": setting.news_post_cap})


# ---------------- Clinical documentation ----------------

@api_view(["GET", "POST"])
def clinical_docs_view(request):
    user = request.user
    if request.method == "POST":
        if user.role != Roles.FIELD_STAFF:
            return Response({"detail": "Only field staff log clinical entries."}, status=403)
        shift_id = request.data.get("shift")
        try:
            shift = Shift.objects.get(id=shift_id, field_staff=user)
        except Shift.DoesNotExist:
            return Response({"detail": "Pick one of your own shifts."}, status=400)
        file = request.FILES.get("file")
        doc = ClinicalDocument.objects.create(
            shift=shift, field_staff=user, client=shift.client,
            notes=request.data.get("notes", ""), file=file, file_name=file.name if file else "",
        )
        return Response(ClinicalDocumentSerializer(doc).data, status=201)
    if user.role == Roles.FIELD_STAFF:
        qs = ClinicalDocument.objects.filter(field_staff=user)
    elif user.role in (Roles.ADMIN, Roles.CUSTOMER_SERVICE, Roles.MANAGER):
        qs = ClinicalDocument.objects.all()
    else:
        qs = ClinicalDocument.objects.none()
    return Response(ClinicalDocumentSerializer(qs.order_by("-created_at"), many=True).data)



@api_view(["POST"])
@permission_classes([IsAdmin])
def programs_bulk_view(request):
    """
    Admin: upload a spreadsheet to create programs in bulk.

    Expected columns, header row required: name, description (description
    is optional). Existing programs are matched by name and have their
    description updated rather than duplicated.
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "Upload an .xlsx file with a file field named file."}, status=400)

    try:
        from openpyxl import load_workbook
    except ImportError:
        return Response({"detail": "openpyxl is not installed on the server."}, status=500)

    try:
        workbook = load_workbook(file, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return Response({"detail": "Could not read that file. Make sure it is a valid .xlsx spreadsheet."}, status=400)

    if not rows:
        return Response({"detail": "The spreadsheet is empty."}, status=400)

    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    if "name" not in header:
        return Response({"detail": "Header row must include: name (description is optional)."}, status=400)
    col = {name: header.index(name) for name in header if name}

    created, updated, errors = [], [], []
    for row_number, row in enumerate(rows[1:], start=2):
        if not row or not any(row):
            continue
        name = str(row[col["name"]] or "").strip() if col.get("name") is not None else ""
        description = str(row[col["description"]] or "").strip() if "description" in col and col["description"] is not None and len(row) > col["description"] else ""
        if not name:
            errors.append({"row": row_number, "reason": "Missing program name."})
            continue
        program, was_created = Program.objects.get_or_create(name=name, defaults={"description": description})
        if was_created:
            created.append({"row": row_number, "name": name})
        else:
            if description:
                program.description = description
                program.save(update_fields=["description"])
            updated.append({"row": row_number, "name": name})

    return Response({"created": created, "updated": updated, "errors": errors})


# ---------------- Manager dashboard ----------------

@api_view(["GET"])
def manager_dashboard_view(request):
    """
    A manager's own dashboard, scoped to the programs assigned to them
    (admins assign programs to a manager the same way they assign programs
    to field staff, from the Users screen). Unlike the CS dashboard, this
    only shows referrals, shifts, and emergencies tied to field staff who
    are in one of the manager's programs, not the whole platform.
    """
    user = request.user
    if user.role != Roles.MANAGER:
        return Response({"detail": "Manager only."}, status=403)

    program_ids = list(user.programs.values_list("id", flat=True))
    staff = User.objects.filter(role=Roles.FIELD_STAFF, programs__id__in=program_ids).distinct()
    staff_ids = list(staff.values_list("id", flat=True))

    referrals = (
        Referral.objects.filter(assigned_staff_id__in=staff_ids)
        .select_related("hospital", "submitted_by", "assigned_staff")
        .order_by("-created_at")
    )
    shifts = (
        Shift.objects.filter(field_staff_id__in=staff_ids)
        .select_related("field_staff", "client")
        .order_by("start_time")
    )
    client_ids = shifts.values_list("client_id", flat=True).distinct()
    emergencies = (
        (EmergencyRequest.objects.filter(reporter_id__in=staff_ids) | EmergencyRequest.objects.filter(client_id__in=client_ids))
        .distinct()
        .select_related("client", "reporter")
        .order_by("-created_at")
    )
    # The approvals queue is still based on direct reports (who actually
    # asks this manager for shift changes), not programs, so it is left
    # exactly as the Approvals page already computes it.
    change_requests = (
        ShiftChangeRequest.objects.filter(manager=user)
        .select_related("shift", "requested_by", "shift__client", "shift__field_staff")
        .order_by("-created_at")
    )

    return Response({
        "programs": ProgramSerializer(Program.objects.filter(id__in=program_ids).order_by("name"), many=True).data,
        "staff_count": staff.count(),
        "referrals": ReferralSerializer(referrals, many=True).data,
        "shifts": ShiftSerializer(shifts, many=True).data,
        "emergencies": EmergencySerializer(emergencies, many=True).data,
        "change_requests": ShiftChangeRequestSerializer(change_requests, many=True).data,
    })
